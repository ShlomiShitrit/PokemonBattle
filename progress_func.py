import openpyxl
from logistics import sprint, check_input


def save_progress(slot, pokemon):
    """
    Save the progress of the game
    :param slot: str
        A slot to save the game in it
    :param pokemon: tuple
        A tuple of the pokemon of the player
    :return: None
    """
    workbook = openpyxl.Workbook()
    save_file = workbook.active
    data = ('Name', 'Attack', 'HP', 'Defense', 'Speed', 'Level', 'XP')
    save_file.append(data)
    for pok in pokemon:
        attributes = (pok.name, pok.atk_pwr, pok.hp, pok.defense, pok.speed, pok.lvl, pok.xp)
        save_file.append(attributes)
    workbook.save(f'slot{slot}.xlsx')


def load_progress():
    """
    Load a saved game
    :return: class openpyxl.worksheet
        An Exel file with the saved game
    """
    flag = True
    while flag:
        what_slot = input('What slot? 1/2/3 ')
        check_what_slot = check_input(what_slot, ('1', '2', '3'))
        try:
            save_file = f'slot{check_what_slot}.xlsx'
            load_save_file = openpyxl.load_workbook(save_file)
        except FileNotFoundError:
            file_error = input('The slot is empty. Are you sure you have a saved game? y/n ')
            check_file_error = check_input(file_error, ('y', 'n'))
            if check_file_error == 'y':
                continue
            else:
                sprint('Starting a new game')
                save_file = 'init_pokemon.xlsx'
                load_save_file = openpyxl.load_workbook(save_file)
                table = load_save_file.active
                flag = False
        else:
            table = load_save_file.active
            flag = False

    return table


def is_save():
    """
    Checks if there is a saved game and if there is one, loads it. otherwise loads a new game
    :return: class openpyxl.worksheet
        An Exel file with the saved game/new game
    """
    save = input('Do you want to load a saved game? y/n ')
    save_check = check_input(save, ('y', 'n'))
    if save_check == 'y':
        table = load_progress()
    else:
        save_file = 'init_pokemon.xlsx'
        load_save_file = openpyxl.load_workbook(save_file)
        table = load_save_file.active
    return table
